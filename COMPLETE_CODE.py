"""
================================================================================
BUSINESS OPERATIONS & FORECASTING ANALYTICS SYSTEM
================================================================================
Complete End-to-End Python Code for:
- KPI Framework Design
- Trend & Seasonality Analysis
- Rolling Metrics
- Demand Forecasting with ML Models
- What-If Scenario Analysis
- Cost vs Revenue Trade-offs
- Customer Churn Analytics
- Workforce Planning
- Regional & Category Analysis
- Executive Summary Dashboard
- Excel Report Generation

Author: Claude (Anthropic)
Date: January 2025
================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# Set plotting style
plt.style.use('seaborn-v0_8-whitegrid')
sns.set_palette("husl")

# =============================================================================
# SECTION 1: DATA GENERATION
# =============================================================================
def generate_realistic_business_data():
    """
    Generate 3 years of realistic synthetic business data with:
    - 5 product categories
    - 5 geographic regions  
    - 4 customer segments
    - Weekly, monthly, and holiday seasonality patterns
    - Year-over-year growth trends
    """
    np.random.seed(42)
    start_date = datetime(2022, 1, 1)
    end_date = datetime(2024, 12, 31)
    dates = pd.date_range(start=start_date, end=end_date, freq='D')
    
    # Category configurations with base revenue, margin, and seasonality
    categories = {
        'Electronics': {'base': 50000, 'margin': 0.25, 'seasonality': 1.3},
        'Apparel': {'base': 35000, 'margin': 0.45, 'seasonality': 1.2},
        'Home & Garden': {'base': 28000, 'margin': 0.35, 'seasonality': 1.4},
        'Food & Beverage': {'base': 42000, 'margin': 0.20, 'seasonality': 1.1},
        'Health & Beauty': {'base': 22000, 'margin': 0.40, 'seasonality': 1.15}
    }
    
    # Regional configurations with multiplier and growth rate
    regions = {
        'North': {'multiplier': 1.2, 'growth': 0.08},
        'South': {'multiplier': 1.0, 'growth': 0.05},
        'East': {'multiplier': 1.15, 'growth': 0.06},
        'West': {'multiplier': 1.3, 'growth': 0.10},
        'Central': {'multiplier': 0.9, 'growth': 0.04}
    }
    
    # Customer segments with weights
    segments = ['Enterprise', 'Mid-Market', 'SMB', 'Consumer']
    segment_weights = [0.15, 0.25, 0.35, 0.25]
    
    records = []
    
    for date in dates:
        day_of_week = date.dayofweek
        month = date.month
        day_of_year = date.timetuple().tm_yday
        year = date.year
        
        # Weekly pattern: weekends 15% lower
        weekly_factor = 1.0 - 0.15 * (day_of_week >= 5)
        
        # Monthly seasonality factors
        monthly_factors = {
            1: 0.75, 2: 0.80, 3: 0.90, 4: 0.95, 5: 1.00, 6: 1.05,
            7: 0.95, 8: 0.90, 9: 1.00, 10: 1.05, 11: 1.30, 12: 1.50
        }
        monthly_factor = monthly_factors[month]
        
        # Years elapsed for growth calculation
        years_elapsed = (date - start_date).days / 365
        
        # Holiday effects
        holiday_factor = 1.0
        if month == 11 and date.day >= 20:  # Black Friday/Thanksgiving
            holiday_factor = 1.8
        elif month == 12 and date.day <= 25:  # Christmas season
            holiday_factor = 1.6 + 0.05 * (25 - date.day)
        elif month == 2 and 10 <= date.day <= 14:  # Valentine's Day
            holiday_factor = 1.2
        elif month == 5 and 10 <= date.day <= 15:  # Mother's Day
            holiday_factor = 1.25
        
        # Generate data for each category-region combination
        for cat_name, cat_props in categories.items():
            for reg_name, reg_props in regions.items():
                # Calculate base revenue with all factors
                base = cat_props['base'] * reg_props['multiplier']
                growth_factor = 1 + reg_props['growth'] * years_elapsed
                seasonal_amplitude = (cat_props['seasonality'] - 1)
                yearly_seasonal = 1 + seasonal_amplitude * np.sin(2 * np.pi * day_of_year / 365 - np.pi/2)
                
                # Final revenue with noise
                revenue = base * weekly_factor * monthly_factor * yearly_seasonal * growth_factor * holiday_factor
                revenue *= np.random.normal(1, 0.08)  # 8% random variation
                
                # Calculate other metrics
                cogs = revenue * (1 - cat_props['margin']) * np.random.normal(1, 0.03)
                gross_profit = revenue - cogs
                marketing_spend = revenue * np.random.uniform(0.05, 0.12)
                operating_expenses = revenue * np.random.uniform(0.08, 0.15)
                avg_order_value = np.random.normal(150, 30) * (1 + 0.1 * years_elapsed)
                orders = int(revenue / avg_order_value)
                customers = int(orders * np.random.uniform(0.7, 0.95))
                new_customers = int(customers * np.random.uniform(0.15, 0.35))
                inventory_level = revenue * np.random.uniform(0.3, 0.6)
                stockout_rate = max(0, np.random.normal(0.03, 0.015))
                employees_needed = max(5, int(revenue / 15000))
                segment = np.random.choice(segments, p=segment_weights)
                
                records.append({
                    'date': date,
                    'year': year,
                    'month': month,
                    'week': date.isocalendar()[1],
                    'day_of_week': day_of_week,
                    'day_name': date.strftime('%A'),
                    'category': cat_name,
                    'region': reg_name,
                    'segment': segment,
                    'revenue': round(revenue, 2),
                    'cogs': round(cogs, 2),
                    'gross_profit': round(gross_profit, 2),
                    'gross_margin': round(gross_profit/revenue, 4),
                    'marketing_spend': round(marketing_spend, 2),
                    'operating_expenses': round(operating_expenses, 2),
                    'net_income': round(gross_profit - marketing_spend - operating_expenses, 2),
                    'orders': orders,
                    'customers': customers,
                    'new_customers': new_customers,
                    'avg_order_value': round(avg_order_value, 2),
                    'inventory_level': round(inventory_level, 2),
                    'stockout_rate': round(stockout_rate, 4),
                    'employees_needed': employees_needed,
                    'holiday_flag': 1 if holiday_factor > 1.1 else 0
                })
    
    return pd.DataFrame(records)


# =============================================================================
# SECTION 2: KPI FRAMEWORK
# =============================================================================
def calculate_kpis(df):
    """Calculate comprehensive KPIs across Financial, Customer, Operational, and Growth dimensions."""
    
    # Aggregate to daily level
    daily_df = df.groupby('date').agg({
        'revenue': 'sum', 'cogs': 'sum', 'gross_profit': 'sum', 'marketing_spend': 'sum',
        'operating_expenses': 'sum', 'net_income': 'sum', 'orders': 'sum', 'customers': 'sum',
        'new_customers': 'sum', 'inventory_level': 'sum', 'employees_needed': 'sum'
    }).reset_index()
    
    kpis = {}
    
    # Financial KPIs
    total_revenue = df['revenue'].sum()
    total_gross_profit = df['gross_profit'].sum()
    total_marketing = df['marketing_spend'].sum()
    total_opex = df['operating_expenses'].sum()
    total_net_income = df['net_income'].sum()
    total_cogs = df['cogs'].sum()
    
    kpis['financial'] = {
        'Total Revenue': f"${total_revenue/1e6:.2f}M",
        'Total Gross Profit': f"${total_gross_profit/1e6:.2f}M",
        'Gross Margin': f"{(total_gross_profit/total_revenue)*100:.1f}%",
        'Net Margin': f"{(total_net_income/total_revenue)*100:.1f}%",
        'EBITDA Margin': f"{((total_net_income + total_opex*0.3)/total_revenue)*100:.1f}%",
        'Marketing ROI': f"{(total_revenue/total_marketing):.2f}x",
        'Operating Expense Ratio': f"{(total_opex/total_revenue)*100:.1f}%"
    }
    
    # Customer KPIs
    total_customers = df['customers'].sum()
    total_new_customers = df['new_customers'].sum()
    total_orders = df['orders'].sum()
    cac = total_marketing / total_new_customers
    ltv = (df['revenue'].sum() / total_customers) * 12
    
    kpis['customer'] = {
        'Total Customers Served': f"{total_customers:,}",
        'New Customer Acquisition': f"{total_new_customers:,}",
        'Customer Acquisition Cost': f"${cac:.2f}",
        'Lifetime Value (Est.)': f"${ltv:.2f}",
        'LTV:CAC Ratio': f"{ltv/cac:.2f}x",
        'Monthly Churn Rate (Est.)': f"5.5%",
        'Orders per Customer': f"{total_orders/total_customers:.2f}"
    }
    
    # Operational KPIs
    avg_inventory = df['inventory_level'].mean()
    inventory_turnover = total_cogs / avg_inventory
    avg_stockout = df['stockout_rate'].mean()
    
    kpis['operational'] = {
        'Average Inventory': f"${avg_inventory/1e6:.2f}M",
        'Inventory Turnover': f"{inventory_turnover:.1f}x annually",
        'Days Inventory Outstanding': f"{365/inventory_turnover:.0f} days",
        'Avg Stockout Rate': f"{avg_stockout*100:.2f}%",
        'Fill Rate': f"{(1-avg_stockout)*100:.2f}%",
        'Avg Daily Orders': f"{total_orders/len(daily_df):,.0f}",
        'Avg Order Value': f"${df['avg_order_value'].mean():.2f}"
    }
    
    # Growth KPIs
    yearly_revenue = df.groupby('year')['revenue'].sum()
    yoy_growth = (yearly_revenue.iloc[-1] - yearly_revenue.iloc[-2]) / yearly_revenue.iloc[-2]
    cagr = (yearly_revenue.iloc[-1] / yearly_revenue.iloc[0]) ** (1/len(yearly_revenue)) - 1
    monthly_revenue = df.groupby(['year', 'month'])['revenue'].sum()
    mom_growth = monthly_revenue.pct_change().mean()
    
    kpis['growth'] = {
        'YoY Revenue Growth': f"{yoy_growth*100:.1f}%",
        'CAGR': f"{cagr*100:.1f}%",
        'Avg MoM Growth': f"{mom_growth*100:.2f}%",
        'Customer Growth Rate': f"{(total_new_customers/total_customers)*100:.1f}%",
        'Revenue per Employee': f"${total_revenue/df['employees_needed'].sum():,.0f}"
    }
    
    return kpis, daily_df


def plot_kpi_dashboard(kpis, df):
    """Create 4-quadrant KPI dashboard visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('KPI Dashboard - Business Performance Overview', fontsize=16, fontweight='bold')
    
    for ax, (title, data, cmap) in zip(axes.flat, [
        ('Financial KPIs', kpis['financial'], 'Blues'),
        ('Customer KPIs', kpis['customer'], 'Greens'),
        ('Operational KPIs', kpis['operational'], 'Oranges'),
        ('Growth KPIs', kpis['growth'], 'Purples')
    ]):
        items = list(data.items())
        colors = plt.cm.get_cmap(cmap)(np.linspace(0.4, 0.8, len(items)))
        ax.barh(range(len(items)), [1]*len(items), color=colors, alpha=0.7)
        for i, (key, val) in enumerate(items):
            ax.text(0.5, i, f"{key}: {val}", va='center', ha='center', fontsize=11, fontweight='bold')
        ax.set_yticks([])
        ax.set_xticks([])
        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlim(0, 1)
    
    plt.tight_layout()
    plt.savefig('/home/claude/01_kpi_dashboard.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   KPI Dashboard saved")


# =============================================================================
# SECTION 3: SEASONALITY ANALYSIS
# =============================================================================
def analyze_seasonality(df):
    """Analyze weekly, monthly, and yearly seasonality patterns."""
    daily_revenue = df.groupby('date')['revenue'].sum().reset_index()
    daily_revenue.set_index('date', inplace=True)
    
    # Weekly aggregation with trend
    weekly_revenue = daily_revenue.resample('W').sum()
    trend = weekly_revenue.rolling(window=13, center=True).mean()
    
    # Monthly seasonality indices
    monthly_revenue = df.groupby('month')['revenue'].sum()
    seasonal_indices = monthly_revenue / monthly_revenue.mean()
    
    # Day-of-week patterns
    dow_revenue = df.groupby('day_name')['revenue'].sum()
    dow_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    dow_revenue = dow_revenue.reindex(dow_order)
    dow_indices = dow_revenue / dow_revenue.mean()
    
    # Year-over-year comparison
    yearly_monthly = df.groupby(['year', 'month'])['revenue'].sum().unstack(level=0)
    
    return {
        'weekly_revenue': weekly_revenue,
        'trend': trend,
        'seasonal_indices': seasonal_indices,
        'dow_indices': dow_indices,
        'yearly_monthly': yearly_monthly
    }


def plot_seasonality_analysis(seasonality_data):
    """Create 4-panel seasonality visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Trend & Seasonality Analysis', fontsize=16, fontweight='bold')
    
    # Weekly revenue with trend
    ax1 = axes[0, 0]
    weekly_rev = seasonality_data['weekly_revenue'] / 1e6
    trend = seasonality_data['trend'] / 1e6
    ax1.plot(weekly_rev.index, weekly_rev.values, alpha=0.5, label='Weekly Revenue')
    ax1.plot(trend.index, trend.values, color='red', linewidth=2.5, label='13-Week Trend')
    ax1.set_ylabel('Revenue ($M)')
    ax1.set_title('Weekly Revenue with Trend Line', fontweight='bold')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Monthly seasonality
    ax2 = axes[0, 1]
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    seasonal = seasonality_data['seasonal_indices'].values
    colors = ['#2ecc71' if s >= 1 else '#e74c3c' for s in seasonal]
    ax2.bar(months, seasonal, color=colors, edgecolor='black', alpha=0.8)
    ax2.axhline(y=1, color='black', linestyle='--', linewidth=1.5)
    ax2.set_ylabel('Seasonal Index')
    ax2.set_title('Monthly Seasonality Index', fontweight='bold')
    ax2.set_ylim(0.6, 1.6)
    
    # Day-of-week pattern
    ax3 = axes[1, 0]
    dow_indices = seasonality_data['dow_indices']
    colors = plt.cm.viridis(np.linspace(0.2, 0.8, 7))
    ax3.bar(dow_indices.index, dow_indices.values, color=colors, edgecolor='black', alpha=0.8)
    ax3.axhline(y=1, color='black', linestyle='--', linewidth=1.5)
    ax3.set_ylabel('Index vs. Average')
    ax3.set_title('Day-of-Week Revenue Pattern', fontweight='bold')
    plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
    
    # YoY comparison
    ax4 = axes[1, 1]
    yearly_monthly = seasonality_data['yearly_monthly'] / 1e6
    for col in yearly_monthly.columns:
        ax4.plot(months, yearly_monthly[col].values, marker='o', linewidth=2, label=f'{col}')
    ax4.set_ylabel('Revenue ($M)')
    ax4.set_title('Year-over-Year Revenue Comparison', fontweight='bold')
    ax4.legend()
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig('/home/claude/02_seasonality_analysis.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Seasonality Analysis saved")


# =============================================================================
# SECTION 4: ROLLING METRICS
# =============================================================================
def calculate_rolling_metrics(df):
    """Calculate rolling averages, growth rates, and volatility metrics."""
    daily = df.groupby('date').agg({
        'revenue': 'sum', 'gross_profit': 'sum', 'orders': 'sum',
        'customers': 'sum', 'new_customers': 'sum'
    }).reset_index()
    daily.set_index('date', inplace=True)
    
    rolling_metrics = pd.DataFrame(index=daily.index)
    rolling_metrics['revenue_7d_avg'] = daily['revenue'].rolling(7).mean()
    rolling_metrics['revenue_30d_avg'] = daily['revenue'].rolling(30).mean()
    rolling_metrics['revenue_30d_std'] = daily['revenue'].rolling(30).std()
    rolling_metrics['revenue_90d_avg'] = daily['revenue'].rolling(90).mean()
    rolling_metrics['revenue_7d_growth'] = daily['revenue'].pct_change(7) * 100
    rolling_metrics['revenue_30d_growth'] = daily['revenue'].pct_change(30) * 100
    rolling_metrics['revenue_cv'] = rolling_metrics['revenue_30d_std'] / rolling_metrics['revenue_30d_avg']
    
    return rolling_metrics, daily


def plot_rolling_metrics(rolling_metrics, daily):
    """Create 4-panel rolling metrics visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Rolling Metrics & Performance Tracking', fontsize=16, fontweight='bold')
    
    # Rolling averages
    ax1 = axes[0, 0]
    ax1.plot(daily.index, daily['revenue']/1e6, alpha=0.3, label='Daily Revenue')
    ax1.plot(rolling_metrics.index, rolling_metrics['revenue_7d_avg']/1e6, linewidth=2, label='7-Day Avg')
    ax1.plot(rolling_metrics.index, rolling_metrics['revenue_30d_avg']/1e6, linewidth=2, label='30-Day Avg')
    ax1.plot(rolling_metrics.index, rolling_metrics['revenue_90d_avg']/1e6, linewidth=2.5, label='90-Day Avg', color='darkred')
    ax1.set_ylabel('Revenue ($M)')
    ax1.set_title('Revenue with Rolling Averages', fontweight='bold')
    ax1.legend(loc='upper left')
    ax1.grid(True, alpha=0.3)
    
    # Growth rates
    ax2 = axes[0, 1]
    ax2.plot(rolling_metrics.index, rolling_metrics['revenue_7d_growth'], alpha=0.5, label='7-Day Growth %')
    ax2.plot(rolling_metrics.index, rolling_metrics['revenue_30d_growth'], linewidth=2, label='30-Day Growth %', color='green')
    ax2.axhline(y=0, color='black', linestyle='--', linewidth=1)
    ax2.set_ylabel('Growth Rate (%)')
    ax2.set_title('Rolling Growth Rates', fontweight='bold')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim(-50, 100)
    
    # Volatility (CV)
    ax3 = axes[1, 0]
    ax3.fill_between(rolling_metrics.index, 0, rolling_metrics['revenue_cv'], alpha=0.5, color='orange')
    ax3.plot(rolling_metrics.index, rolling_metrics['revenue_cv'], color='darkorange', linewidth=1)
    ax3.set_ylabel('Coefficient of Variation')
    ax3.set_title('Revenue Volatility (30-Day Rolling CV)', fontweight='bold')
    ax3.grid(True, alpha=0.3)
    
    # Confidence bands
    ax4 = axes[1, 1]
    upper_band = rolling_metrics['revenue_30d_avg'] + 2*rolling_metrics['revenue_30d_std']
    lower_band = rolling_metrics['revenue_30d_avg'] - 2*rolling_metrics['revenue_30d_std']
    ax4.fill_between(rolling_metrics.index, lower_band/1e6, upper_band/1e6, alpha=0.3, color='blue', label='2-sigma Band')
    ax4.plot(rolling_metrics.index, rolling_metrics['revenue_30d_avg']/1e6, linewidth=2, color='blue', label='30-Day Avg')
    ax4.plot(daily.index, daily['revenue']/1e6, alpha=0.5, color='gray', label='Daily')
    ax4.set_ylabel('Revenue ($M)')
    ax4.set_title('Revenue with Confidence Bands', fontweight='bold')
    ax4.legend()
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig('/home/claude/03_rolling_metrics.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Rolling Metrics saved")


# =============================================================================
# SECTION 5: FORECASTING MODELS
# =============================================================================
def build_forecasting_models(df):
    """Build and compare 4 ML models for revenue forecasting."""
    
    # Prepare daily data
    daily = df.groupby('date').agg({
        'revenue': 'sum', 'orders': 'sum', 'marketing_spend': 'sum'
    }).reset_index()
    daily.set_index('date', inplace=True)
    
    # Feature engineering
    daily['day_of_week'] = daily.index.dayofweek
    daily['month'] = daily.index.month
    daily['day_of_year'] = daily.index.dayofyear
    daily['week_of_year'] = daily.index.isocalendar().week
    daily['quarter'] = daily.index.quarter
    daily['is_weekend'] = (daily['day_of_week'] >= 5).astype(int)
    
    # Lag features
    for lag in [1, 7, 14, 30]:
        daily[f'revenue_lag_{lag}'] = daily['revenue'].shift(lag)
    
    # Rolling averages
    daily['revenue_7d_ma'] = daily['revenue'].rolling(7).mean()
    daily['revenue_30d_ma'] = daily['revenue'].rolling(30).mean()
    
    # Cyclical encoding
    daily['month_sin'] = np.sin(2 * np.pi * daily['month'] / 12)
    daily['month_cos'] = np.cos(2 * np.pi * daily['month'] / 12)
    daily['dow_sin'] = np.sin(2 * np.pi * daily['day_of_week'] / 7)
    daily['dow_cos'] = np.cos(2 * np.pi * daily['day_of_week'] / 7)
    
    # Clean data
    daily_clean = daily.dropna()
    
    # Feature columns
    feature_cols = [
        'day_of_week', 'month', 'day_of_year', 'week_of_year', 'quarter', 'is_weekend',
        'revenue_lag_1', 'revenue_lag_7', 'revenue_lag_14', 'revenue_lag_30',
        'revenue_7d_ma', 'revenue_30d_ma', 'month_sin', 'month_cos', 'dow_sin', 'dow_cos',
        'marketing_spend'
    ]
    
    # Train/test split (last 90 days for testing)
    X = daily_clean[feature_cols]
    y = daily_clean['revenue']
    train_size = len(X) - 90
    X_train, X_test = X[:train_size], X[train_size:]
    y_train, y_test = y[:train_size], y[train_size:]
    
    # Scale features for linear models
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    # Define models
    models = {
        'Linear Regression': LinearRegression(),
        'Ridge Regression': Ridge(alpha=1.0),
        'Random Forest': RandomForestRegressor(n_estimators=100, max_depth=10, random_state=42),
        'Gradient Boosting': GradientBoostingRegressor(n_estimators=100, max_depth=5, random_state=42)
    }
    
    # Train and evaluate
    results = {}
    predictions = {}
    
    for name, model in models.items():
        if 'Forest' in name or 'Boosting' in name:
            model.fit(X_train, y_train)
            pred = model.predict(X_test)
        else:
            model.fit(X_train_scaled, y_train)
            pred = model.predict(X_test_scaled)
        
        results[name] = {
            'MAE': mean_absolute_error(y_test, pred),
            'RMSE': np.sqrt(mean_squared_error(y_test, pred)),
            'R2': r2_score(y_test, pred),
            'MAPE': np.mean(np.abs((y_test - pred) / y_test)) * 100
        }
        predictions[name] = pred
    
    # Select best model
    best_model_name = min(results, key=lambda x: results[x]['MAPE'])
    best_predictions = predictions[best_model_name]
    
    # Calculate confidence intervals
    residuals = y_test.values - best_predictions
    residual_std = np.std(residuals)
    
    return {
        'results': results,
        'predictions': predictions,
        'y_test': y_test,
        'y_train': y_train,
        'best_model': best_model_name,
        'ci_lower': best_predictions - 1.96 * residual_std,
        'ci_upper': best_predictions + 1.96 * residual_std,
        'residual_std': residual_std,
        'models': models,
        'scaler': scaler,
        'feature_cols': feature_cols,
        'daily': daily_clean
    }


def plot_forecasting_results(forecast_data):
    """Create 4-panel forecasting results visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Demand/Revenue Forecasting Results', fontsize=16, fontweight='bold')
    
    y_test = forecast_data['y_test']
    predictions = forecast_data['predictions']
    results = forecast_data['results']
    
    # Model comparison
    ax1 = axes[0, 0]
    model_names = list(results.keys())
    mape_values = [results[m]['MAPE'] for m in model_names]
    colors = ['#2ecc71' if v == min(mape_values) else '#3498db' for v in mape_values]
    bars = ax1.barh(model_names, mape_values, color=colors, edgecolor='black')
    ax1.set_xlabel('MAPE (%)')
    ax1.set_title('Model Comparison (Lower is Better)', fontweight='bold')
    for bar, val in zip(bars, mape_values):
        ax1.text(val + 0.1, bar.get_y() + bar.get_height()/2, f'{val:.2f}%', va='center')
    
    # Actual vs predicted
    ax2 = axes[0, 1]
    best_model = forecast_data['best_model']
    best_pred = predictions[best_model]
    ci_lower = forecast_data['ci_lower']
    ci_upper = forecast_data['ci_upper']
    ax2.fill_between(range(len(y_test)), ci_lower/1e6, ci_upper/1e6, alpha=0.3, color='blue', label='95% CI')
    ax2.plot(range(len(y_test)), y_test.values/1e6, 'o-', markersize=3, label='Actual', alpha=0.7)
    ax2.plot(range(len(y_test)), best_pred/1e6, 's-', markersize=3, label=f'Predicted ({best_model})', alpha=0.7)
    ax2.set_xlabel('Days')
    ax2.set_ylabel('Revenue ($M)')
    ax2.set_title('Actual vs Predicted with 95% CI', fontweight='bold')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    # Residual distribution
    ax3 = axes[1, 0]
    residuals = y_test.values - best_pred
    ax3.hist(residuals/1e6, bins=30, edgecolor='black', alpha=0.7, color='purple')
    ax3.axvline(x=0, color='red', linestyle='--', linewidth=2)
    ax3.set_xlabel('Residual ($M)')
    ax3.set_ylabel('Frequency')
    ax3.set_title('Residual Distribution', fontweight='bold')
    
    # Performance table
    ax4 = axes[1, 1]
    ax4.axis('off')
    table_data = [[name, f"${r['MAE']/1e6:.3f}M", f"${r['RMSE']/1e6:.3f}M", f"{r['R2']:.3f}", f"{r['MAPE']:.2f}%"]
                  for name, r in results.items()]
    table = ax4.table(cellText=table_data, colLabels=['Model', 'MAE', 'RMSE', 'R2', 'MAPE'], 
                      loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1.2, 1.8)
    for i, name in enumerate(model_names):
        if name == best_model:
            for j in range(5):
                table[(i+1, j)].set_facecolor('#90EE90')
    ax4.set_title('Model Performance Metrics', fontweight='bold', pad=20)
    
    plt.tight_layout()
    plt.savefig('/home/claude/04_forecasting_results.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Forecasting Results saved")


# =============================================================================
# SECTION 6: FUTURE FORECAST
# =============================================================================
def generate_future_forecast(forecast_data, df, periods=90):
    """Generate N-day forward forecast with confidence intervals."""
    daily = forecast_data['daily']
    best_model_name = forecast_data['best_model']
    models = forecast_data['models']
    scaler = forecast_data['scaler']
    feature_cols = forecast_data['feature_cols']
    
    last_date = daily.index[-1]
    future_dates = pd.date_range(start=last_date + timedelta(days=1), periods=periods, freq='D')
    
    future_data = []
    recent_revenues = list(daily['revenue'].tail(30))
    
    for i, date in enumerate(future_dates):
        row = {
            'day_of_week': date.dayofweek,
            'month': date.month,
            'day_of_year': date.timetuple().tm_yday,
            'week_of_year': date.isocalendar()[1],
            'quarter': (date.month - 1) // 3 + 1,
            'is_weekend': 1 if date.dayofweek >= 5 else 0,
            'month_sin': np.sin(2 * np.pi * date.month / 12),
            'month_cos': np.cos(2 * np.pi * date.month / 12),
            'dow_sin': np.sin(2 * np.pi * date.dayofweek / 7),
            'dow_cos': np.cos(2 * np.pi * date.dayofweek / 7),
            'marketing_spend': daily['marketing_spend'].mean()
        }
        
        # Lag features from recent predictions
        row['revenue_lag_1'] = recent_revenues[-1] if len(recent_revenues) >= 1 else daily['revenue'].iloc[-1]
        row['revenue_lag_7'] = recent_revenues[-7] if len(recent_revenues) >= 7 else daily['revenue'].mean()
        row['revenue_lag_14'] = recent_revenues[-14] if len(recent_revenues) >= 14 else daily['revenue'].mean()
        row['revenue_lag_30'] = recent_revenues[-30] if len(recent_revenues) >= 30 else daily['revenue'].mean()
        row['revenue_7d_ma'] = np.mean(recent_revenues[-7:]) if len(recent_revenues) >= 7 else daily['revenue'].mean()
        row['revenue_30d_ma'] = np.mean(recent_revenues[-30:]) if len(recent_revenues) >= 30 else daily['revenue'].mean()
        
        future_data.append(row)
        
        # Predict and update recent revenues
        X_pred = pd.DataFrame([row])[feature_cols]
        model = models[best_model_name]
        if 'Forest' in best_model_name or 'Boosting' in best_model_name:
            pred = model.predict(X_pred)[0]
        else:
            pred = model.predict(scaler.transform(X_pred))[0]
        
        recent_revenues.append(pred)
        if len(recent_revenues) > 30:
            recent_revenues.pop(0)
    
    # Final predictions
    future_df = pd.DataFrame(future_data, index=future_dates)
    X_future = future_df[feature_cols]
    model = models[best_model_name]
    
    if 'Forest' in best_model_name or 'Boosting' in best_model_name:
        future_predictions = model.predict(X_future)
    else:
        future_predictions = model.predict(scaler.transform(X_future))
    
    residual_std = forecast_data['residual_std']
    
    return {
        'dates': future_dates,
        'predictions': future_predictions,
        'ci_lower': future_predictions - 1.96 * residual_std,
        'ci_upper': future_predictions + 1.96 * residual_std
    }


def plot_future_forecast(forecast_data, future_forecast, periods=90):
    """Create 2-panel future forecast visualization."""
    fig, axes = plt.subplots(2, 1, figsize=(16, 10))
    fig.suptitle(f'{periods}-Day Revenue Forecast', fontsize=16, fontweight='bold')
    
    y_train = forecast_data['y_train']
    y_test = forecast_data['y_test']
    historical_dates = list(y_train.index) + list(y_test.index)
    historical_values = list(y_train.values) + list(y_test.values)
    
    # Historical + forecast
    ax1 = axes[0]
    ax1.plot(historical_dates[-180:], np.array(historical_values[-180:])/1e6, 'b-', alpha=0.7, linewidth=1, label='Historical')
    ax1.fill_between(future_forecast['dates'], future_forecast['ci_lower']/1e6, future_forecast['ci_upper']/1e6, 
                     alpha=0.3, color='green', label='95% CI')
    ax1.plot(future_forecast['dates'], future_forecast['predictions']/1e6, 'g-', linewidth=2, label='Forecast')
    ax1.axvline(x=historical_dates[-1], color='red', linestyle='--', linewidth=2, label='Forecast Start')
    ax1.set_ylabel('Revenue ($M)')
    ax1.set_title('Historical + Forecast View', fontweight='bold')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Monthly summary
    ax2 = axes[1]
    future_df = pd.DataFrame({
        'date': future_forecast['dates'],
        'forecast': future_forecast['predictions'],
        'ci_lower': future_forecast['ci_lower'],
        'ci_upper': future_forecast['ci_upper']
    })
    future_df['month'] = future_df['date'].dt.to_period('M')
    monthly_forecast = future_df.groupby('month').agg({'forecast': 'sum', 'ci_lower': 'sum', 'ci_upper': 'sum'})
    
    months = [str(m) for m in monthly_forecast.index]
    forecasts = monthly_forecast['forecast'].values / 1e6
    ci_low = monthly_forecast['ci_lower'].values / 1e6
    ci_high = monthly_forecast['ci_upper'].values / 1e6
    
    x = np.arange(len(months))
    ax2.bar(x, forecasts, color='#3498db', edgecolor='black', alpha=0.8, label='Forecast')
    ax2.errorbar(x, forecasts, yerr=[forecasts-ci_low, ci_high-forecasts], fmt='none', color='black', capsize=5, label='95% CI')
    ax2.set_xticks(x)
    ax2.set_xticklabels(months, rotation=45)
    ax2.set_ylabel('Revenue ($M)')
    ax2.set_title('Monthly Forecast Summary', fontweight='bold')
    ax2.legend()
    ax2.grid(True, alpha=0.3, axis='y')
    
    for i, v in enumerate(forecasts):
        ax2.text(i, v + 1, f'${v:.1f}M', ha='center', fontsize=10, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig('/home/claude/05_future_forecast.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Future Forecast saved")


# =============================================================================
# SECTION 7: SCENARIO ANALYSIS
# =============================================================================
def scenario_analysis(df, forecast_data):
    """Run what-if scenario analysis with 7 different business scenarios."""
    total_revenue = df['revenue'].sum()
    total_marketing = df['marketing_spend'].sum()
    total_profit = df['net_income'].sum()
    
    scenarios = {
        'Base Case': {
            'marketing_change': 0, 'price_change': 0, 'cost_reduction': 0, 'demand_shock': 0,
            'description': 'Current business'
        },
        'Marketing Boost (+30%)': {
            'marketing_change': 0.30, 'price_change': 0, 'cost_reduction': 0, 'demand_shock': 0.15,
            'description': 'Aggressive marketing'
        },
        'Price Increase (+10%)': {
            'marketing_change': 0, 'price_change': 0.10, 'cost_reduction': 0, 'demand_shock': -0.05,
            'description': 'Premium pricing'
        },
        'Cost Optimization (-15%)': {
            'marketing_change': -0.10, 'price_change': 0, 'cost_reduction': 0.15, 'demand_shock': -0.02,
            'description': 'Efficiency focus'
        },
        'Recession Scenario': {
            'marketing_change': -0.20, 'price_change': -0.05, 'cost_reduction': 0.10, 'demand_shock': -0.20,
            'description': 'Economic downturn'
        },
        'Expansion Strategy': {
            'marketing_change': 0.50, 'price_change': -0.05, 'cost_reduction': -0.05, 'demand_shock': 0.25,
            'description': 'Growth push'
        },
        'Holiday Peak': {
            'marketing_change': 0.50, 'price_change': 0.05, 'cost_reduction': 0, 'demand_shock': 0.35,
            'description': 'Holiday optimization'
        }
    }
    
    results = []
    for name, params in scenarios.items():
        new_revenue = total_revenue * (1 + params['demand_shock'] + params['price_change'])
        new_marketing = total_marketing * (1 + params['marketing_change'])
        new_cogs = df['cogs'].sum() * (1 + params['demand_shock']) * (1 - params['cost_reduction'])
        new_gross_profit = new_revenue - new_cogs
        new_opex = df['operating_expenses'].sum() * (1 - params['cost_reduction'] * 0.5)
        new_net_income = new_gross_profit - new_marketing - new_opex
        
        results.append({
            'Scenario': name,
            'Description': params['description'],
            'Revenue': new_revenue,
            'Revenue Change': (new_revenue / total_revenue - 1) * 100,
            'Marketing Spend': new_marketing,
            'Gross Profit': new_gross_profit,
            'Net Income': new_net_income,
            'Net Income Change': (new_net_income / total_profit - 1) * 100,
            'Gross Margin': new_gross_profit / new_revenue * 100,
            'Net Margin': new_net_income / new_revenue * 100,
            'ROI': (new_net_income / new_marketing) * 100
        })
    
    return pd.DataFrame(results)


def plot_scenario_analysis(scenario_df):
    """Create 4-panel scenario analysis visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('What-If Scenario Analysis', fontsize=16, fontweight='bold')
    scenarios = scenario_df['Scenario'].values
    
    # Revenue impact
    ax1 = axes[0, 0]
    revenue_change = scenario_df['Revenue Change'].values
    sorted_idx = np.argsort(revenue_change)
    colors = plt.cm.RdYlGn(np.linspace(0.2, 0.8, len(scenarios)))
    ax1.barh([scenarios[i] for i in sorted_idx], [revenue_change[i] for i in sorted_idx], 
             color=[colors[i] for i in sorted_idx], edgecolor='black')
    ax1.axvline(x=0, color='black', linestyle='-', linewidth=1)
    ax1.set_xlabel('Revenue Change (%)')
    ax1.set_title('Revenue Impact by Scenario', fontweight='bold')
    ax1.grid(True, alpha=0.3, axis='x')
    
    # Net income impact
    ax2 = axes[0, 1]
    ni_change = scenario_df['Net Income Change'].values
    sorted_idx = np.argsort(ni_change)
    colors = ['#2ecc71' if ni_change[i] >= 0 else '#e74c3c' for i in sorted_idx]
    ax2.barh([scenarios[i] for i in sorted_idx], [ni_change[i] for i in sorted_idx], 
             color=colors, edgecolor='black')
    ax2.axvline(x=0, color='black', linestyle='-', linewidth=1)
    ax2.set_xlabel('Net Income Change (%)')
    ax2.set_title('Profitability Impact by Scenario', fontweight='bold')
    ax2.grid(True, alpha=0.3, axis='x')
    
    # Margin comparison
    ax3 = axes[1, 0]
    x = np.arange(len(scenarios))
    width = 0.35
    ax3.bar(x - width/2, scenario_df['Gross Margin'].values, width, label='Gross Margin', color='#3498db', edgecolor='black')
    ax3.bar(x + width/2, scenario_df['Net Margin'].values, width, label='Net Margin', color='#2ecc71', edgecolor='black')
    ax3.set_xticks(x)
    ax3.set_xticklabels(scenarios, rotation=45, ha='right')
    ax3.set_ylabel('Margin (%)')
    ax3.set_title('Margin Comparison Across Scenarios', fontweight='bold')
    ax3.legend()
    ax3.grid(True, alpha=0.3, axis='y')
    
    # ROI
    ax4 = axes[1, 1]
    roi = scenario_df['ROI'].values
    colors = plt.cm.coolwarm(np.linspace(0.2, 0.8, len(scenarios)))
    ax4.bar(scenarios, roi, color=colors, edgecolor='black')
    ax4.set_xticklabels(scenarios, rotation=45, ha='right')
    ax4.set_ylabel('ROI (%)')
    ax4.set_title('Return on Investment by Scenario', fontweight='bold')
    ax4.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.savefig('/home/claude/06_scenario_analysis.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Scenario Analysis saved")


# =============================================================================
# SECTION 8: COST-REVENUE ANALYSIS
# =============================================================================
def cost_revenue_analysis(df):
    """Analyze cost structure and break-even points."""
    monthly = df.groupby(['year', 'month']).agg({
        'revenue': 'sum', 'cogs': 'sum', 'gross_profit': 'sum', 'marketing_spend': 'sum',
        'operating_expenses': 'sum', 'net_income': 'sum', 'orders': 'sum', 
        'customers': 'sum', 'new_customers': 'sum'
    }).reset_index()
    
    monthly['cogs_ratio'] = monthly['cogs'] / monthly['revenue']
    monthly['marketing_ratio'] = monthly['marketing_spend'] / monthly['revenue']
    monthly['opex_ratio'] = monthly['operating_expenses'] / monthly['revenue']
    monthly['gross_margin'] = monthly['gross_profit'] / monthly['revenue']
    monthly['net_margin'] = monthly['net_income'] / monthly['revenue']
    
    # Break-even calculation
    fixed_costs = monthly['operating_expenses'].mean() * 0.7
    variable_cost_ratio = monthly['cogs_ratio'].mean()
    contribution_margin = 1 - variable_cost_ratio
    break_even_revenue = fixed_costs / contribution_margin
    
    return {
        'monthly': monthly,
        'break_even_revenue': break_even_revenue,
        'fixed_costs': fixed_costs,
        'contribution_margin': contribution_margin
    }


def plot_cost_revenue_analysis(analysis_data):
    """Create 4-panel cost-revenue analysis visualization."""
    monthly = analysis_data['monthly']
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Cost vs Revenue Trade-off Analysis', fontsize=16, fontweight='bold')
    
    # Cost structure
    ax1 = axes[0, 0]
    ax1.stackplot(range(len(monthly)), 
                  monthly['cogs_ratio']*100, monthly['marketing_ratio']*100, monthly['opex_ratio']*100,
                  labels=['COGS', 'Marketing', 'OpEx'], colors=['#e74c3c', '#3498db', '#f1c40f'], alpha=0.8)
    ax1.set_xlabel('Month')
    ax1.set_ylabel('% of Revenue')
    ax1.set_title('Cost Structure Over Time', fontweight='bold')
    ax1.legend(loc='upper right')
    ax1.set_ylim(0, 100)
    ax1.grid(True, alpha=0.3)
    
    # Margin trends
    ax2 = axes[0, 1]
    ax2.plot(range(len(monthly)), monthly['gross_margin']*100, 'o-', linewidth=2, markersize=4, label='Gross Margin')
    ax2.plot(range(len(monthly)), monthly['net_margin']*100, 's-', linewidth=2, markersize=4, label='Net Margin')
    ax2.set_xlabel('Month')
    ax2.set_ylabel('Margin (%)')
    ax2.set_title('Margin Trends Over Time', fontweight='bold')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    # Marketing vs revenue
    ax3 = axes[1, 0]
    ax3.scatter(monthly['marketing_spend']/1e6, monthly['revenue']/1e6, 
                c=monthly.index, cmap='viridis', s=100, alpha=0.7, edgecolor='black')
    z = np.polyfit(monthly['marketing_spend'], monthly['revenue'], 1)
    p = np.poly1d(z)
    x_line = np.linspace(monthly['marketing_spend'].min(), monthly['marketing_spend'].max(), 100)
    ax3.plot(x_line/1e6, p(x_line)/1e6, 'r--', linewidth=2, label=f'Trend (slope={z[0]:.2f})')
    ax3.set_xlabel('Marketing Spend ($M)')
    ax3.set_ylabel('Revenue ($M)')
    ax3.set_title('Marketing Spend vs Revenue', fontweight='bold')
    ax3.legend()
    ax3.grid(True, alpha=0.3)
    
    # Break-even
    ax4 = axes[1, 1]
    revenues = np.linspace(0, monthly['revenue'].max() * 1.2, 100)
    fixed_costs = analysis_data['fixed_costs']
    variable_costs = revenues * (1 - analysis_data['contribution_margin'])
    total_costs = fixed_costs + variable_costs
    ax4.plot(revenues/1e6, revenues/1e6, 'g-', linewidth=2, label='Revenue')
    ax4.plot(revenues/1e6, total_costs/1e6, 'r-', linewidth=2, label='Total Costs')
    ax4.fill_between(revenues/1e6, total_costs/1e6, revenues/1e6, 
                     where=(revenues >= total_costs), alpha=0.3, color='green', label='Profit Zone')
    ax4.fill_between(revenues/1e6, total_costs/1e6, revenues/1e6, 
                     where=(revenues < total_costs), alpha=0.3, color='red', label='Loss Zone')
    ax4.axvline(x=analysis_data['break_even_revenue']/1e6, color='black', linestyle='--', linewidth=2, 
                label=f'Break-even: ${analysis_data["break_even_revenue"]/1e6:.1f}M')
    ax4.set_xlabel('Revenue ($M)')
    ax4.set_ylabel('Value ($M)')
    ax4.set_title('Break-Even Analysis', fontweight='bold')
    ax4.legend()
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig('/home/claude/07_cost_revenue_analysis.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Cost-Revenue Analysis saved")


# =============================================================================
# SECTION 9: CUSTOMER CHURN ANALYSIS
# =============================================================================
def customer_churn_analysis(df):
    """Analyze customer churn with cohort analysis and segment-level metrics."""
    np.random.seed(42)
    
    # Generate cohort data
    cohorts = pd.date_range(start='2022-01', end='2024-10', freq='MS')
    cohort_data = []
    
    for cohort_date in cohorts:
        initial_customers = np.random.randint(800, 1500)
        for month_offset in range(min(36, (datetime(2024, 12, 31) - cohort_date.to_pydatetime()).days // 30)):
            retention_rate = 0.95 * np.exp(-0.03 * month_offset) + 0.05 * np.random.normal(0, 0.02)
            retention_rate = max(0.1, min(1.0, retention_rate))
            remaining_customers = int(initial_customers * retention_rate ** month_offset)
            cohort_data.append({
                'cohort': cohort_date,
                'month_number': month_offset,
                'customers': remaining_customers,
                'initial_customers': initial_customers,
                'retention_rate': remaining_customers / initial_customers
            })
    
    cohort_df = pd.DataFrame(cohort_data)
    avg_retention_by_month = cohort_df.groupby('month_number')['retention_rate'].mean()
    
    # Segment-level metrics
    segment_metrics = df.groupby('segment').agg({
        'revenue': 'sum', 'customers': 'sum', 'new_customers': 'sum'
    }).reset_index()
    
    churn_rates = {'Enterprise': 0.03, 'Mid-Market': 0.05, 'SMB': 0.08, 'Consumer': 0.12}
    segment_metrics['churn_rate'] = segment_metrics['segment'].map(churn_rates)
    segment_metrics['lifetime_months'] = 1 / segment_metrics['churn_rate']
    segment_metrics['revenue_per_customer'] = segment_metrics['revenue'] / segment_metrics['customers']
    segment_metrics['ltv'] = segment_metrics['revenue_per_customer'] * segment_metrics['lifetime_months']
    
    return {
        'cohort_df': cohort_df,
        'avg_retention': avg_retention_by_month,
        'segment_metrics': segment_metrics
    }


def plot_churn_analysis(churn_data):
    """Create 4-panel churn analysis visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Customer Churn & Retention Analytics', fontsize=16, fontweight='bold')
    cohort_df = churn_data['cohort_df']
    segment_metrics = churn_data['segment_metrics']
    
    # Cohort heatmap
    ax1 = axes[0, 0]
    pivot_table = cohort_df.pivot_table(values='retention_rate', index='cohort', columns='month_number', aggfunc='mean')
    pivot_subset = pivot_table.iloc[-12:, :12]
    sns.heatmap(pivot_subset * 100, annot=True, fmt='.0f', cmap='RdYlGn', ax=ax1, 
                vmin=0, vmax=100, cbar_kws={'label': 'Retention %'})
    ax1.set_xlabel('Month Since Acquisition')
    ax1.set_ylabel('Cohort')
    ax1.set_title('Cohort Retention Heatmap', fontweight='bold')
    
    # Retention curve
    ax2 = axes[0, 1]
    avg_retention = churn_data['avg_retention']
    ax2.plot(avg_retention.index, avg_retention.values * 100, 'o-', linewidth=2, markersize=6, color='#3498db')
    ax2.fill_between(avg_retention.index, 0, avg_retention.values * 100, alpha=0.3, color='#3498db')
    ax2.set_xlabel('Months Since Acquisition')
    ax2.set_ylabel('Retention Rate (%)')
    ax2.set_title('Average Retention Curve', fontweight='bold')
    ax2.set_ylim(0, 105)
    ax2.grid(True, alpha=0.3)
    
    # Churn by segment
    ax3 = axes[1, 0]
    x = np.arange(len(segment_metrics))
    width = 0.35
    ax3.bar(x - width/2, segment_metrics['churn_rate'] * 100, width, 
            label='Monthly Churn %', color='#e74c3c', edgecolor='black')
    ax3.bar(x + width/2, segment_metrics['lifetime_months'], width, 
            label='Avg Lifetime (months)', color='#2ecc71', edgecolor='black')
    ax3.set_xticks(x)
    ax3.set_xticklabels(segment_metrics['segment'])
    ax3.set_ylabel('Value')
    ax3.set_title('Churn Rate & Lifetime by Segment', fontweight='bold')
    ax3.legend()
    ax3.grid(True, alpha=0.3, axis='y')
    
    # LTV by segment
    ax4 = axes[1, 1]
    colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(segment_metrics)))
    bars = ax4.bar(segment_metrics['segment'], segment_metrics['ltv']/1000, color=colors, edgecolor='black')
    ax4.set_ylabel('Customer LTV ($K)')
    ax4.set_title('Customer Lifetime Value by Segment', fontweight='bold')
    ax4.grid(True, alpha=0.3, axis='y')
    for bar, val in zip(bars, segment_metrics['ltv']/1000):
        ax4.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, 
                 f'${val:.1f}K', ha='center', fontsize=11, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig('/home/claude/08_churn_analysis.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Churn Analysis saved")


# =============================================================================
# SECTION 10: WORKFORCE PLANNING
# =============================================================================
def workforce_planning_analysis(df, forecast_data, future_forecast):
    """Analyze workforce requirements with scenario planning."""
    monthly_workforce = df.groupby(['year', 'month']).agg({
        'revenue': 'sum', 'orders': 'sum', 'employees_needed': 'sum'
    }).reset_index()
    
    monthly_workforce['revenue_per_employee'] = monthly_workforce['revenue'] / monthly_workforce['employees_needed']
    monthly_workforce['orders_per_employee'] = monthly_workforce['orders'] / monthly_workforce['employees_needed']
    avg_rev_per_employee = monthly_workforce['revenue_per_employee'].mean()
    
    monthly_factors = monthly_workforce.groupby('month')['employees_needed'].mean()
    avg_factor = monthly_factors.mean()
    seasonal_factors = monthly_factors / avg_factor
    
    scenarios = {
        'Baseline': {'productivity_change': 0, 'demand_change': 0},
        'High Demand (+20%)': {'productivity_change': 0, 'demand_change': 0.20},
        'Automation (+15%)': {'productivity_change': 0.15, 'demand_change': 0},
        'Recession (-15%)': {'productivity_change': 0, 'demand_change': -0.15},
        'Expansion': {'productivity_change': 0.10, 'demand_change': 0.25}
    }
    
    current_employees = monthly_workforce['employees_needed'].iloc[-1]
    forecast_revenue = future_forecast['predictions']
    
    scenario_results = []
    for name, params in scenarios.items():
        adjusted_demand = forecast_revenue.mean() * (1 + params['demand_change'])
        adjusted_productivity = avg_rev_per_employee * (1 + params['productivity_change'])
        needed_employees = adjusted_demand / adjusted_productivity
        scenario_results.append({
            'Scenario': name,
            'Employees Needed': int(needed_employees),
            'Change vs Current': int(needed_employees - current_employees),
            'Change %': (needed_employees / current_employees - 1) * 100
        })
    
    return {
        'monthly_workforce': monthly_workforce,
        'seasonal_factors': seasonal_factors,
        'scenario_results': pd.DataFrame(scenario_results)
    }


def plot_workforce_analysis(workforce_data):
    """Create 4-panel workforce planning visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Workforce Planning', fontsize=16, fontweight='bold')
    monthly = workforce_data['monthly_workforce']
    
    # Historical staffing
    ax1 = axes[0, 0]
    ax1.bar(range(len(monthly)), monthly['employees_needed'], color='#3498db', alpha=0.8)
    ax1.set_xlabel('Month')
    ax1.set_ylabel('Employees')
    ax1.set_title('Historical Staffing', fontweight='bold')
    
    # Productivity
    ax2 = axes[0, 1]
    ax2.plot(range(len(monthly)), monthly['revenue_per_employee']/1000, 'o-', label='Rev/Emp ($K)')
    ax2.set_title('Productivity', fontweight='bold')
    ax2.legend()
    
    # Seasonal factors
    ax3 = axes[1, 0]
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    ax3.bar(months, workforce_data['seasonal_factors'].values)
    ax3.axhline(y=1, linestyle='--')
    ax3.set_title('Seasonal Factors', fontweight='bold')
    
    # Scenarios
    ax4 = axes[1, 1]
    scenarios = workforce_data['scenario_results']
    ax4.barh(scenarios['Scenario'], scenarios['Employees Needed'])
    ax4.set_title('Workforce Scenarios', fontweight='bold')
    
    plt.tight_layout()
    plt.savefig('/home/claude/09_workforce_planning.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Workforce Planning saved")


# =============================================================================
# SECTION 11: REGIONAL & CATEGORY ANALYSIS
# =============================================================================
def regional_category_analysis(df):
    """Analyze performance by region and category."""
    region_metrics = df.groupby('region').agg({
        'revenue': 'sum', 'gross_profit': 'sum', 'marketing_spend': 'sum'
    }).reset_index()
    region_metrics['gross_margin'] = region_metrics['gross_profit'] / region_metrics['revenue']
    region_metrics['marketing_roi'] = region_metrics['revenue'] / region_metrics['marketing_spend']
    
    category_metrics = df.groupby('category').agg({
        'revenue': 'sum', 'gross_profit': 'sum', 'inventory_level': 'mean'
    }).reset_index()
    category_metrics['gross_margin'] = category_metrics['gross_profit'] / category_metrics['revenue']
    category_metrics['inventory_turnover'] = category_metrics['revenue'] / category_metrics['inventory_level']
    
    cross_analysis = df.groupby(['region', 'category'])['revenue'].sum().unstack()
    yearly_region = df.groupby(['year', 'region'])['revenue'].sum().unstack()
    region_growth = (yearly_region.iloc[-1] - yearly_region.iloc[-2]) / yearly_region.iloc[-2] * 100
    
    return {
        'region_metrics': region_metrics,
        'category_metrics': category_metrics,
        'cross_analysis': cross_analysis,
        'region_growth': region_growth
    }


def plot_regional_category(analysis_data):
    """Create 4-panel regional and category analysis visualization."""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Regional & Category Analysis', fontsize=16, fontweight='bold')
    
    # Revenue by region pie
    ax1 = axes[0, 0]
    region_metrics = analysis_data['region_metrics']
    ax1.pie(region_metrics['revenue'], labels=region_metrics['region'], autopct='%1.1f%%')
    ax1.set_title('Revenue by Region', fontweight='bold')
    
    # Category margins
    ax2 = axes[0, 1]
    category_metrics = analysis_data['category_metrics']
    x = np.arange(len(category_metrics))
    ax2.bar(x, category_metrics['gross_margin']*100, label='Margin %')
    ax2.set_xticks(x)
    ax2.set_xticklabels(category_metrics['category'], rotation=45, ha='right')
    ax2.set_title('Category Margins', fontweight='bold')
    
    # Cross-analysis heatmap
    ax3 = axes[1, 0]
    cross = analysis_data['cross_analysis'] / 1e6
    sns.heatmap(cross, annot=True, fmt='.1f', cmap='YlOrRd', ax=ax3)
    ax3.set_title('Revenue Heatmap ($M)', fontweight='bold')
    
    # Regional growth
    ax4 = axes[1, 1]
    region_growth = analysis_data['region_growth']
    colors = ['#2ecc71' if g >= 0 else '#e74c3c' for g in region_growth.values]
    ax4.bar(region_growth.index, region_growth.values, color=colors)
    ax4.set_title('YoY Growth by Region', fontweight='bold')
    
    plt.tight_layout()
    plt.savefig('/home/claude/10_regional_category_analysis.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Regional Analysis saved")


# =============================================================================
# SECTION 12: EXECUTIVE SUMMARY DASHBOARD
# =============================================================================
def create_executive_summary(kpis, scenario_df, forecast_data, future_forecast):
    """Create one-page executive summary dashboard."""
    fig = plt.figure(figsize=(20, 14))
    fig.suptitle('EXECUTIVE SUMMARY DASHBOARD', fontsize=20, fontweight='bold')
    gs = fig.add_gridspec(3, 4, hspace=0.3, wspace=0.3)
    
    # KPI cards
    for i, (title, value, color) in enumerate([
        ('Total Revenue', kpis['financial']['Total Revenue'], '#2ecc71'),
        ('Gross Margin', kpis['financial']['Gross Margin'], '#3498db'),
        ('LTV:CAC', kpis['customer']['LTV:CAC Ratio'], '#9b59b6'),
        ('YoY Growth', kpis['growth']['YoY Revenue Growth'], '#e74c3c')
    ]):
        ax = fig.add_subplot(gs[0, i])
        ax.text(0.5, 0.6, value, fontsize=22, ha='center', fontweight='bold', color=color)
        ax.text(0.5, 0.25, title, fontsize=11, ha='center')
        ax.axis('off')
    
    # Forecast chart
    ax5 = fig.add_subplot(gs[1, :2])
    y_test = forecast_data['y_test']
    ax5.plot(range(60), y_test.values[-60:]/1e6, 'b-', label='Historical')
    ax5.plot(range(60, 60+90), future_forecast['predictions']/1e6, 'g-', label='Forecast')
    ax5.fill_between(range(60, 60+90), future_forecast['ci_lower']/1e6, 
                     future_forecast['ci_upper']/1e6, alpha=0.3, color='green')
    ax5.axvline(x=60, color='red', linestyle='--')
    ax5.set_title('90-Day Forecast', fontweight='bold')
    ax5.legend()
    
    # Scenario impact
    ax6 = fig.add_subplot(gs[1, 2:])
    ni_change = scenario_df['Net Income Change'].values[:5]
    colors = ['#2ecc71' if c >= 0 else '#e74c3c' for c in ni_change]
    ax6.barh(scenario_df['Scenario'].values[:5], ni_change, color=colors)
    ax6.axvline(x=0, color='black')
    ax6.set_title('Scenario Impact', fontweight='bold')
    
    # Key insights
    ax7 = fig.add_subplot(gs[2, :])
    ax7.axis('off')
    insights = [f"Best Model: {forecast_data['best_model']} | Revenue: {kpis['financial']['Total Revenue']} | Marketing ROI: {kpis['financial']['Marketing ROI']}"]
    ax7.text(0.5, 0.5, insights[0], fontsize=14, ha='center', transform=ax7.transAxes)
    
    plt.savefig('/home/claude/11_executive_summary.png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print("   Executive Summary saved")


# =============================================================================
# SECTION 13: EXCEL REPORT GENERATION
# =============================================================================
def generate_excel_report(df, kpis, scenario_df, forecast_data, future_forecast, 
                         seasonality_data, churn_data, workforce_data, regional_data):
    """Generate comprehensive Excel report with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2C3E50')
    
    # Executive Summary sheet
    ws = wb.active
    ws.title = "Executive Summary"
    ws['A1'] = "BUSINESS OPERATIONS ANALYTICS"
    ws['A1'].font = Font(bold=True, size=18)
    
    row = 3
    for section, data in [('FINANCIAL', kpis['financial']), ('CUSTOMER', kpis['customer']),
                          ('OPERATIONAL', kpis['operational']), ('GROWTH', kpis['growth'])]:
        ws[f'A{row}'] = section
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for key, val in data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = val
            row += 1
        row += 1
    
    # Scenarios sheet
    ws2 = wb.create_sheet("Scenarios")
    for col, header in enumerate(scenario_df.columns, 1):
        ws2.cell(row=1, column=col, value=header).font = header_font
        ws2.cell(row=1, column=col).fill = header_fill
    for row_idx, row_data in enumerate(scenario_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Forecast sheet
    ws3 = wb.create_sheet("Forecast")
    for col, h in enumerate(['Date', 'Forecast', 'CI_Low', 'CI_High'], 1):
        ws3.cell(row=1, column=col, value=h).font = header_font
        ws3.cell(row=1, column=col).fill = header_fill
    for i, (d, p, l, h) in enumerate(zip(future_forecast['dates'], future_forecast['predictions'],
                                          future_forecast['ci_lower'], future_forecast['ci_upper']), 2):
        ws3.cell(row=i, column=1, value=d.strftime('%Y-%m-%d'))
        ws3.cell(row=i, column=2, value=p)
        ws3.cell(row=i, column=3, value=l)
        ws3.cell(row=i, column=4, value=h)
    
    output_path = '/home/claude/business_analytics_report.xlsx'
    wb.save(output_path)
    print(f"   Excel Report saved")
    return output_path


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    """Main execution function."""
    print("="*80)
    print("BUSINESS OPERATIONS & FORECASTING ANALYTICS SYSTEM")
    print("="*80)
    print()
    
    # Step 1: Generate Data
    print("Step 1: Generating realistic business data...")
    df = generate_realistic_business_data()
    print(f"   Generated {len(df):,} records from {df['date'].min().strftime('%Y-%m-%d')} to {df['date'].max().strftime('%Y-%m-%d')}")
    print(f"   Categories: {df['category'].nunique()} | Regions: {df['region'].nunique()} | Segments: {df['segment'].nunique()}")
    
    # Step 2: KPIs
    print("\nStep 2: Building KPI Framework...")
    kpis, daily_df = calculate_kpis(df)
    plot_kpi_dashboard(kpis, df)
    
    # Step 3: Seasonality
    print("\nStep 3: Analyzing Seasonality & Trends...")
    seasonality_data = analyze_seasonality(df)
    plot_seasonality_analysis(seasonality_data)
    
    # Step 4: Rolling Metrics
    print("\nStep 4: Computing Rolling Metrics...")
    rolling_metrics, daily = calculate_rolling_metrics(df)
    plot_rolling_metrics(rolling_metrics, daily)
    
    # Step 5: Forecasting
    print("\nStep 5: Building Forecasting Models...")
    forecast_data = build_forecasting_models(df)
    plot_forecasting_results(forecast_data)
    print(f"   Best Model: {forecast_data['best_model']} (MAPE: {forecast_data['results'][forecast_data['best_model']]['MAPE']:.2f}%)")
    
    # Step 6: Future Forecast
    print("\nStep 6: Generating 90-Day Forecast...")
    future_forecast = generate_future_forecast(forecast_data, df, periods=90)
    plot_future_forecast(forecast_data, future_forecast, periods=90)
    print(f"   Forecast Total: ${sum(future_forecast['predictions'])/1e6:.2f}M")
    
    # Step 7: Scenarios
    print("\nStep 7: Running What-If Scenario Analysis...")
    scenario_df = scenario_analysis(df, forecast_data)
    plot_scenario_analysis(scenario_df)
    
    # Step 8: Cost-Revenue
    print("\nStep 8: Analyzing Cost vs Revenue Trade-offs...")
    cost_analysis = cost_revenue_analysis(df)
    plot_cost_revenue_analysis(cost_analysis)
    print(f"   Break-even Revenue: ${cost_analysis['break_even_revenue']/1e6:.2f}M monthly")
    
    # Step 9: Churn
    print("\nStep 9: Customer Churn Analytics...")
    churn_data = customer_churn_analysis(df)
    plot_churn_analysis(churn_data)
    
    # Step 10: Workforce
    print("\nStep 10: Workforce Planning Analysis...")
    workforce_data = workforce_planning_analysis(df, forecast_data, future_forecast)
    plot_workforce_analysis(workforce_data)
    
    # Step 11: Regional
    print("\nStep 11: Regional & Category Breakdown...")
    regional_data = regional_category_analysis(df)
    plot_regional_category(regional_data)
    
    # Step 12: Executive Summary
    print("\nStep 12: Creating Executive Summary Dashboard...")
    create_executive_summary(kpis, scenario_df, forecast_data, future_forecast)
    
    # Step 13: Excel Report
    print("\nStep 13: Generating Excel Report...")
    generate_excel_report(df, kpis, scenario_df, forecast_data, future_forecast,
                         seasonality_data, churn_data, workforce_data, regional_data)
    
    # Step 14: Save Data
    print("\nStep 14: Saving Raw Data...")
    df.to_csv('/home/claude/business_data.csv', index=False)
    print("   Raw data saved: business_data.csv")
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE!")
    print("="*80)
    print("\nGenerated Outputs:")
    print("  01_kpi_dashboard.png")
    print("  02_seasonality_analysis.png")
    print("  03_rolling_metrics.png")
    print("  04_forecasting_results.png")
    print("  05_future_forecast.png")
    print("  06_scenario_analysis.png")
    print("  07_cost_revenue_analysis.png")
    print("  08_churn_analysis.png")
    print("  09_workforce_planning.png")
    print("  10_regional_category_analysis.png")
    print("  11_executive_summary.png")
    print("  business_analytics_report.xlsx")
    print("  business_data.csv")
    
    return df, kpis, scenario_df, forecast_data, future_forecast


# =============================================================================
# RUN
# =============================================================================
if __name__ == "__main__":
    df, kpis, scenario_df, forecast_data, future_forecast = main()
